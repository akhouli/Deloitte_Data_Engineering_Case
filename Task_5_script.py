import pandas as pd
import os
from openpyxl import Workbook

# Directory containing CSV files
data_dir = os.path.expanduser("Case_Study_Data_For_Share")

# Initialize lists to hold dataframes and inconsistency reports
dataframes = []
inconsistency_reports = []

# Function to extract data from a single CSV file
def extract_data(file_path):
    try:
        df = pd.read_csv(file_path, delimiter='|')
        print(f"Successfully read {file_path}")
        return df
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None

# Function to profile data for inconsistencies
def profile_data(df):
    inconsistencies = {}

    required_columns = [
        'Order ID', 'Order Date', 'Ship Date', 'Customer ID', 'Customer Name',
        'Sales', 'Quantity', 'Discount', 'Profit', 'Postal Code', 'Country'
    ]
    
    # Check for missing columns
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        inconsistencies['Missing Columns'] = missing_columns

    if 'Order Date' in df.columns:
        # Check for invalid data formats
        invalid_dates = pd.to_datetime(df['Order Date'], errors='coerce').isna().sum()
        if invalid_dates > 0:
            inconsistencies['Invalid Data Formats'] = ['Order Date']

    if 'Sales' in df.columns and 'Quantity' in df.columns:
        # Check for records with zero sales and zero quantity
        zero_sales_quantity = df[(df['Sales'] == 0) & (df['Quantity'] == 0)].index.tolist()
        if zero_sales_quantity:
            inconsistencies['Zero Sales and Quantity'] = zero_sales_quantity

        # Check for negative sales values
        negative_sales = df[df['Sales'] < 0].index.tolist()
        if negative_sales:
            inconsistencies['Negative Sales Values'] = negative_sales

    if 'Discount' in df.columns:
        # Check for unrealistic discount values
        invalid_discounts = df[(df['Discount'] < 0) | (df['Discount'] > 1)].index.tolist()
        if invalid_discounts:
            inconsistencies['Unrealistic Discount Values'] = invalid_discounts

    if 'Postal Code' in df.columns:
        # Check for invalid postal codes (assuming US postal codes here for simplicity)
        invalid_postal_codes = df[~df['Postal Code'].astype(str).str.match(r'^\d{5}(-\d{4})?$')].index.tolist()
        if invalid_postal_codes:
            inconsistencies['Invalid Postal Codes'] = invalid_postal_codes

    if 'Country' in df.columns:
        # Check for inconsistent country names
        valid_countries = ['United States']
        invalid_countries = df[~df['Country'].isin(valid_countries)].index.tolist()
        if invalid_countries:
            inconsistencies['Inconsistent Country Names'] = invalid_countries

    if 'Order Date' in df.columns and 'Ship Date' in df.columns:
        # Check for mismatched order and ship dates
        mismatched_dates = df[df['Ship Date'] < df['Order Date']].index.tolist()
        if mismatched_dates:
            inconsistencies['Mismatched Order and Ship Dates'] = mismatched_dates

    if 'Customer ID' in df.columns and 'Customer Name' in df.columns:
        # Check for inconsistent Customer IDs (ensure same customer name has the same customer ID)
        inconsistent_customer_ids = df.groupby('Customer Name')['Customer ID'].nunique()
        inconsistent_customer_ids = inconsistent_customer_ids[inconsistent_customer_ids > 1].index.tolist()
        inconsistent_customer_ids_rows = df[df['Customer Name'].isin(inconsistent_customer_ids)].index.tolist()
        if inconsistent_customer_ids_rows:
            inconsistencies['Inconsistent Customer IDs'] = inconsistent_customer_ids_rows

    if 'Profit' in df.columns:
        # Check for unrealistic profit values (negative profit)
        negative_profits = df[df['Profit'] < 0].index.tolist()
        if negative_profits:
            inconsistencies['Negative Profit Values'] = negative_profits

    return inconsistencies

# Extract and profile data from all files
for file_name in os.listdir(data_dir):
    if file_name.endswith(".csv"):
        file_path = os.path.join(data_dir, file_name)
        df = extract_data(file_path)
        if df is not None:
            dataframes.append(df)
            inconsistencies = profile_data(df)
            if inconsistencies:
                inconsistency_reports.append({'file': file_name, 'inconsistencies': inconsistencies, 'data': df})

# Check if any dataframes were loaded
if not dataframes:
    print("No dataframes were loaded. Please check the file paths and formats.")
else:
    # Combine all dataframes into one
    all_data = pd.concat(dataframes, ignore_index=True)

    # Data Cleansing Function
    def clean_data(df):
        # Remove duplicates
        df.drop_duplicates(inplace=True)

        # Handle missing values
        df['Customer Name'].fillna('Unknown', inplace=True)
        df.dropna(subset=['Order ID', 'Product ID', 'Customer ID', 'Order Date', 'Sales'], inplace=True)

        # Correct data types
        df['Order Date'] = pd.to_datetime(df['Order Date'], errors='coerce')
        df['Ship Date'] = pd.to_datetime(df['Ship Date'], errors='coerce')
        df['Sales'] = pd.to_numeric(df['Sales'], errors='coerce')
        df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce')
        df['Discount'] = pd.to_numeric(df['Discount'], errors='coerce')
        df['Profit'] = pd.to_numeric(df['Profit'], errors='coerce')

        return df

    # Clean the combined data
    cleaned_data = clean_data(all_data)

    # Generate Inconsistency Report in Excel

    # Create a new Excel workbook
    wb = Workbook()

    # Inconsistencies Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Inconsistencies_Summary"
    ws_summary.append(["Inconsistency Type", "Description", "Suggestion to handle", "Distinct Count of Row ID"])

    # Examples of Inconsistencies Sheet
    ws_examples = wb.create_sheet(title="Inconsistencies_Examples")
    ws_examples.append(["Inconsistency Type", "Row ID", "Order ID", "Order Date", "Ship Date", "Ship Mode", "Customer ID", "Customer Name", "Segment", "Country", "City", "State", "Postal Code", "Region", "Product ID", "Category", "Sub-Category", "Product Name", "Sales", "Quantity", "Discount", "Profit"])

    # Data Quality Report Sheet
    ws_quality = wb.create_sheet(title="Quality_Report")
    ws_quality.append(["Inconsistency Type", "File Name", "Row ID", "Description"])

    # Generate Inconsistency Report
    summary_dict = {}
    for report in inconsistency_reports:
        file_name = report['file']
        inconsistencies = report['inconsistencies']
        df = report['data']

        for inc_type, cols in inconsistencies.items():
            if inc_type not in summary_dict:
                summary_dict[inc_type] = {"Description": "", "Suggestion to handle": "", "Distinct Count of Row ID": 0}

            if inc_type == "Zero Sales and Quantity":
                description = "Zero sales and zero quantity found"
                suggestion = "Handle programmatically: Correct or remove records with zero sales and zero quantity"
                row_count = len(cols)
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for zero sales and quantity
                example_rows = df.loc[cols].head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])
            elif inc_type == "Missing Columns":
                description = f"Inconsistency found in columns: {', '.join(cols)}"
                suggestion = "Requires SME input: Standardize data types across columns"
                row_count = df.shape[0]
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for missing columns
                example_rows = df.head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])
            elif inc_type == "Negative Sales Values":
                description = "Negative sales values found"
                suggestion = "Handle programmatically: Remove or correct negative sales values"
                row_count = len(cols)
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for negative sales values
                example_rows = df.loc[cols].head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])
            elif inc_type == "Unrealistic Discount Values":
                description = "Unrealistic discount values found"
                suggestion = "Handle programmatically: Ensure discount values are between 0 and 1"
                row_count = len(cols)
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for unrealistic discount values
                example_rows = df.loc[cols].head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])
            elif inc_type == "Invalid Postal Codes":
                description = "Invalid postal codes found"
                suggestion = "Handle programmatically: Ensure postal codes follow the correct format"
                row_count = len(cols)
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for invalid postal codes
                example_rows = df.loc[cols].head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])
            elif inc_type == "Inconsistent Country Names":
                description = "Inconsistent country names found"
                suggestion = "Handle programmatically: Ensure country names match the predefined list"
                row_count = len(cols)
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for inconsistent country names
                example_rows = df.loc[cols].head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])
            elif inc_type == "Mismatched Order and Ship Dates":
                description = "Mismatched order and ship dates found"
                suggestion = "Investigate Further: Ensure ship dates are not before order dates"
                row_count = len(cols)
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for mismatched order and ship dates
                example_rows = df.loc[cols].head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])
            elif inc_type == "Inconsistent Customer IDs":
                description = "Inconsistent customer IDs found"
                suggestion = "Handle programmatically: Ensure customer IDs are consistent for the same customer name"
                row_count = len(cols)
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for inconsistent customer IDs
                example_rows = df.loc[cols].head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])
            elif inc_type == "Negative Profit Values":
                description = "Negative profit values found"
                suggestion = "Requires Business input : Investigate reasons for negative profit"
                row_count = len(cols)
                summary_dict[inc_type]["Description"] = description
                summary_dict[inc_type]["Suggestion to handle"] = suggestion
                summary_dict[inc_type]["Distinct Count of Row ID"] += row_count

                # Add examples for negative profit values
                example_rows = df.loc[cols].head(2).to_dict('records')
                for row in example_rows:
                    ws_examples.append([inc_type] + [row.get(col, "") for col in df.columns])

    # Populate Inconsistencies_Summary sheet
    for inc_type, details in summary_dict.items():
        if inc_type == "Missing Columns" and details["Distinct Count of Row ID"] == 0:
            continue  # Skip missing columns with zero count
        ws_summary.append([inc_type, details["Description"], details["Suggestion to handle"], details["Distinct Count of Row ID"]])

    # Generate Data Quality Report
    for report in inconsistency_reports:
        file_name = report['file']
        inconsistencies = report['inconsistencies']
        df = report['data']

        for inc_type, cols in inconsistencies.items():
            description = summary_dict[inc_type]["Description"]
            if inc_type in ["Zero Sales and Quantity", "Negative Sales Values", "Unrealistic Discount Values", "Invalid Postal Codes", "Inconsistent Country Names", "Mismatched Order and Ship Dates", "Inconsistent Customer IDs", "Negative Profit Values"]:
                for row_id in cols:
                    ws_quality.append([inc_type, file_name, row_id, description])
            else:
                for row_id in df.head(2).index:
                    ws_quality.append([inc_type, file_name, row_id, description])

    # Save workbook
    wb.save("Task_5_Inconsistencies_Analysis.xlsx")

    print("Task_5_Inconsistencies_Analysis.xlsx has been created successfully.")
